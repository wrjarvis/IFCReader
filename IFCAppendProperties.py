from IFCReader import IfcFile
from openpyxl import load_workbook
import random


def read_append_data(excel_filename):
    print('Reading append excel file')
    workbook = load_workbook(excel_filename, data_only=True)
    append_dict = {}
    for worksheet in workbook.worksheets:
        property_set = worksheet.title
        append_dict[property_set] = {}
        for row_i in range(2, worksheet.max_row + 1):
            append_dict[property_set][worksheet.cell(row_i, 1).value] = {}
            for col_i in range(2, worksheet.max_column + 1):
                append_dict[property_set][worksheet.cell(row_i, 1).value][worksheet.cell(1, col_i).value] =\
                    worksheet.cell(row_i, col_i).value
    return append_dict


def create_single_property(ifc_hash, name, value):
    if isinstance(value, int):
        out_label = 'IFCINTEGER(' + str(value) + ')'
    elif isinstance(value, float):
        out_label = 'IFCREAL(' + str(value) + ')'
    else:
        out_label = "IFCLABEL('" + str(value) + "')"
    return "#" + str(ifc_hash) + "= IFCPROPERTYSINGLEVALUE('" + str(name) + "',$," + out_label + ",$);"


def create_guid():
    lower = 'abcdefghijklmnopqrstyvwxyz'
    upper = lower.upper()
    numbers = '1234567890'
    full_string = lower + upper + numbers
    guid = ''
    for i in range(0, 22):
        guid += random.choice(full_string)
    return guid


def create_property_set(set_name, property_list, owner_hash, ifc_hash):
    property_string = '(' + ','.join(['#' + str(x) for x in property_list])
    guid = create_guid()
    return "#" + str(ifc_hash) + "= IFCPROPERTYSET('" + guid + "'," + owner_hash + ",'" + str(set_name) + "','" +\
           str(set_name) + "'," + property_string + "));"


def create_assignment(ifc_hash, owner_hash, set_hash, item_hash):
    guid = create_guid()
    return "#" + str(ifc_hash) + "= IFCRELDEFINESBYPROPERTIES('" + guid + "'," + owner_hash + ",$,$,(" + item_hash +\
           "),#" + str(set_hash) + ");"


def process_append_data(append_data, ifc_file):
    print('Processing append data')
    hash_no = max([int(x[1:]) for x in ifc_file.Entities.keys()]) + 1
    owner_hash = list(ifc_file.EntitiesByType['IFCOWNERHISTORY'].keys())[0]
    new_ifc_data = []
    for set_name in append_data.keys():
        for item_guid in append_data[set_name].keys():
            item_id = ifc_file.GlobalID[item_guid]
            property_list = []
            for property_name in append_data[set_name][item_guid]:
                property_list.append(hash_no)
                new_ifc_data.append(create_single_property(hash_no, property_name,
                                                           append_data[set_name][item_guid][property_name]))
                hash_no += 1
            new_ifc_data.append(create_property_set(set_name, property_list, owner_hash, hash_no))
            set_hash = hash_no
            hash_no += 1
            new_ifc_data.append(create_assignment(hash_no, owner_hash, set_hash, item_id))
            hash_no += 1
    return new_ifc_data


def create_new_ifc(ifc_file, new_ifc_data, overwrite):
    print('Creating new IFC')
    new_ifc = []
    data_section = False
    for line in ifc_file.File:
        if line == 'DATA;':
            data_section = True
        elif line == 'ENDSEC;' and data_section:
            new_ifc += new_ifc_data
            data_section = False
        if overwrite:
            if 'IFCPROPERTYSINGLEVALUE' in line or 'IFCPROPERTYSET' in line or 'IFCRELDEFINESBYPROPERTIES' in line:
                pass
            else:
                new_ifc.append(line)
    return new_ifc


def write_ifc(new_ifc, ifc_file_name):
    print('Saving new IFC')
    f = open(ifc_file_name, 'w')
    for line in new_ifc:
        f.write(line + '\n')


def append_properties(ifc_file, append_excel, append_ifc_filename, overwrite):
    append_data = read_append_data(append_excel)
    new_ifc_data = process_append_data(append_data, ifc_file)
    new_ifc = create_new_ifc(ifc_file, new_ifc_data, overwrite)
    write_ifc(new_ifc, append_ifc_filename)


if __name__ == '__main__':
    IfcFile = IfcFile("test_data/AC9R1-Haus-G-H-Ver2-2x3.ifc", "Schema/IFC2X3_TC1.exp")
    AppendExcelFileName = 'AppendData.xlsx'
    AppendIFCFileName = 'NewIFC.ifc'
    append_properties(IfcFile, AppendExcelFileName, AppendIFCFileName, False)
